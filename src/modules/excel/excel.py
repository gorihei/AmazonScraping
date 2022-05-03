from openpyxl import Workbook, load_workbook

class Excel:
    
    alphabet = {
        1 : 'A',
        2 : 'B',
        3 : 'C',
        4 : 'D',
        5 : 'E',
        6 : 'F',
        7 : 'G',
        8 : 'H',
        9 : 'I',
        10 : 'J',
        11 : 'K',
        12 : 'L',
        13 : 'M',
        14 : 'N',
        15 : 'O',
        16 : 'P',
        17 : 'Q',
        18 : 'R',
        19 : 'S',
        20 : 'T',
        21 : 'U',
        22 : 'V',
        23 : 'W',
        24 : 'X',
        25 : 'Y',
        26 : 'Z'
    }
    
    def __init__(self, path):
        self.path = path
        self.wb = load_workbook(path)
        
    def __del__(self):
        if self.wb != None:
            self.wb.close()
            
    def get_sheet_names(self):
        return self.book.sheetnames

    def get_book(self):
        return self.book

    def get_sheet(self, name):
        return self.wb[name]
    
    def add_sheet(self, name):
        return self.wb.create_sheet(name)
    
    def remove_sheet(self, name):
        self.book.remove(name)
        
    def save(self, path=''):
        if path == '':
            path = self.path
        self.wb.save(path)
        
    def get_row_values(self, ws, idx):
        values = [list(filter(None,list(cell))) for cell in ws.iter_cols(min_col=ws.min_column, max_col=ws.max_column, min_row=idx, max_row=idx, values_only=True)]
        return [item for value in values for item in value]
    
    def change_col_number_from_index(self, idx):
        return self.alphabet[idx]
    
    
    